import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(10)

driver.get("https://www.orangehrm.com/en/book-a-free-demo")
FullName = driver.find_element(By.ID, "Form_getForm_FullName")
PhoneNumber = driver.find_element(By.ID, "Form_getForm_Contact")
Email = driver.find_element(By.ID, "Form_getForm_Email")
CompanyName = driver.find_element(By.ID, "Form_getForm_CompanyName")
Country = driver.find_element(By.ID, "Form_getForm_Country")
NoEmp = driver.find_element(By.ID, "Form_getForm_NoOfEmployees")

workbook = openpyxl.load_workbook("SeleniumExcelDDT.xlsx")
#sheet_name= workbook["login"]
sheet_name= workbook["registration"]


''''#Note : ❌ Problem:
In openpyxl, rows and columns are 1-indexed (they start from 1, not 0).
So you cannot use 0 as the column index.'''

#get total number of rows
row_count = sheet_name.max_row
print(row_count)
col_count= sheet_name.max_column
print(col_count)

for curr_row in range(2, row_count+1):
    name= sheet_name.cell(curr_row, 1).value
    number= sheet_name.cell(curr_row, 2).value
    email= sheet_name.cell(curr_row, 3).value
    company= sheet_name.cell(curr_row, 4).value
    country= sheet_name.cell(curr_row, 5).value
    EmployeeNumber= sheet_name.cell(curr_row, 6).value

    print(str(name) + str(number) + str(email) + str(company) + str(country) + str(EmployeeNumber))

    FullName.clear()
    FullName.send_keys(str(name))
    PhoneNumber.clear()
    PhoneNumber.send_keys(str(number))
    Email.clear()
    Email.send_keys(str(email))
    CompanyName.clear()
    CompanyName.send_keys(str(company))

    Country.send_keys(str(country))

    NoEmp.send_keys(str(EmployeeNumber))

time.sleep(4)

  #  print(str(user_name) + "" +str(password))


input("Exit")
